"""Export service for Telegram Attendance Bot.

Provides functionality to generate daily reports, monthly Excel exports, and CSV reports.
"""

from __future__ import annotations

import csv
import io
from calendar import monthrange
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from typing import Dict, List, Optional, Tuple

import pytz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func

from src.config import get_config
from src.database import (
    AttendanceLog,
    AttendanceType,
    PointLog,
    User,
    UserStatus,
    get_db_session,
)


def _get_is_late(log, fallback: bool = False) -> bool:
    """Safely read is_late flag from AttendanceLog (older rows may not have attribute)."""
    return getattr(log, "is_late", fallback)


@dataclass
class DailyReportData:
    """Data class for daily attendance report."""

    date: date
    total_employees: int
    checked_in: int
    on_time: int
    late: int
    not_checked_in: int
    checked_out: int
    present_users: List[Tuple[str, datetime, bool]]  # (name, check_in_time, is_late)
    absent_users: List[str]
    late_users: List[Tuple[str, datetime, int]]  # (name, check_in_time, late_minutes)


@dataclass
class EmployeeMonthlyData:
    """Data class for individual employee monthly attendance data."""

    user_id: int
    full_name: str
    daily_records: Dict[int, Tuple[Optional[datetime], Optional[datetime], bool]]  # day -> (in, out, is_late)
    total_days_present: int = 0
    late_days: int = 0
    total_late_minutes: int = 0


@dataclass
class MonthlyReportData:
    """Data class for monthly attendance report."""

    year: int
    month: int
    employee_data: List[EmployeeMonthlyData] = field(default_factory=list)
    summary: Dict[str, int] = field(default_factory=dict)


@dataclass
class MonthlyPointRow:
    """Điểm tháng theo từng nguồn của một user."""
    
    user_id: int
    full_name: str
    total_points: int
    meeting_points: int
    evidence_points: int
    penalty_points: int
    absence_points: int
    other_points: int


class ExportService:
    """Service for exporting attendance data to various formats."""

    # Excel styles as class variables
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    LATE_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    ABSENT_FILL = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @staticmethod
    def get_daily_report(target_date: Optional[date] = None) -> DailyReportData:
        """
        Get daily attendance report data.

        Args:
            target_date: The date to generate the report for. Defaults to today.

        Returns:
            DailyReportData containing attendance statistics for the day.
        """
        config = get_config()
        tz = pytz.timezone(config.timezone.timezone)

        if target_date is None:
            target_date = datetime.now(tz).date()

        # Define the day boundaries in local timezone
        day_start = tz.localize(datetime.combine(target_date, time.min))
        day_end = tz.localize(datetime.combine(target_date, time.max))

        # Work start time for late calculation
        work_start = tz.localize(
            datetime.combine(
                target_date,
                time(config.attendance.work_start_hour, config.attendance.work_start_minute),
            )
        )
        late_threshold = work_start + timedelta(minutes=config.attendance.late_threshold_minutes)

        with get_db_session() as session:
            # Get all active employees
            active_users = (
                session.query(User)
                .filter(User.status == UserStatus.ACTIVE)
                .all()
            )
            total_employees = len(active_users)
            user_map = {user.user_id: user.full_name for user in active_users}

            # Get all check-ins for the day
            check_ins = (
                session.query(AttendanceLog)
                .filter(
                    and_(
                        AttendanceLog.type == AttendanceType.IN,
                        AttendanceLog.timestamp >= day_start,
                        AttendanceLog.timestamp <= day_end,
                    )
                )
                .all()
            )

            # Get all check-outs for the day
            check_outs = (
                session.query(AttendanceLog)
                .filter(
                    and_(
                        AttendanceLog.type == AttendanceType.OUT,
                        AttendanceLog.timestamp >= day_start,
                        AttendanceLog.timestamp <= day_end,
                    )
                )
                .all()
            )

            # Build check-in data (earliest check-in per user)
            user_check_ins: Dict[int, AttendanceLog] = {}
            for log in check_ins:
                if log.user_id not in user_check_ins or log.timestamp < user_check_ins[log.user_id].timestamp:
                    user_check_ins[log.user_id] = log

            # Build check-out data (latest check-out per user)
            user_check_outs: Dict[int, AttendanceLog] = {}
            for log in check_outs:
                if log.user_id not in user_check_outs or log.timestamp > user_check_outs[log.user_id].timestamp:
                    user_check_outs[log.user_id] = log

            # Calculate statistics
            checked_in = len(user_check_ins)
            checked_out = len(user_check_outs)

            present_users: List[Tuple[str, datetime, bool]] = []
            late_users: List[Tuple[str, datetime, int]] = []
            on_time = 0
            late = 0

            for user_id, log in user_check_ins.items():
                user_name = user_map.get(user_id, f"User {user_id}")
                check_in_time = log.timestamp
                # Ensure timezone-aware for comparison
                check_in_local = (
                    check_in_time.astimezone(tz)
                    if check_in_time.tzinfo
                    else tz.localize(check_in_time)
                )
                is_late = _get_is_late(log) or check_in_local > late_threshold

                present_users.append((user_name, check_in_local, is_late))

                if is_late:
                    late += 1
                    late_minutes = int((check_in_local - work_start).total_seconds() / 60)
                    late_users.append((user_name, check_in_local, max(0, late_minutes)))
                else:
                    on_time += 1

            # Calculate absent users
            checked_in_user_ids = set(user_check_ins.keys())
            absent_users = [
                user_map[user_id]
                for user_id in user_map.keys()
                if user_id not in checked_in_user_ids
            ]
            not_checked_in = len(absent_users)

            # Sort lists
            present_users.sort(key=lambda x: x[1])  # Sort by check-in time
            late_users.sort(key=lambda x: x[2], reverse=True)  # Sort by late minutes descending
            absent_users.sort()

            return DailyReportData(
                date=target_date,
                total_employees=total_employees,
                checked_in=checked_in,
                on_time=on_time,
                late=late,
                not_checked_in=not_checked_in,
                checked_out=checked_out,
                present_users=present_users,
                absent_users=absent_users,
                late_users=late_users,
            )

    @staticmethod
    def format_daily_report(report: DailyReportData) -> str:
        """
        Format daily report data as a text string for Telegram message.

        Args:
            report: DailyReportData to format.

        Returns:
            Formatted string suitable for Telegram message.
        """
        config = get_config()
        tz = pytz.timezone(config.timezone.timezone)

        lines = [
            f"BÁO CÁO CHẤM CÔNG NGÀY {report.date.strftime('%d/%m/%Y')}",
            "=" * 40,
            "",
            "TỔNG QUAN:",
            f"  Tổng nhân viên: {report.total_employees}",
            f"  Đã check-in: {report.checked_in}",
            f"  Đúng giờ: {report.on_time}",
            f"  Đi muộn: {report.late}",
            f"  Chưa check-in: {report.not_checked_in}",
            f"  Đã check-out: {report.checked_out}",
            "",
        ]

        if report.late_users:
            lines.append("NHÂN VIÊN ĐI MUỘN:")
            for name, check_in_time, late_minutes in report.late_users:
                local_time = check_in_time.astimezone(tz) if check_in_time.tzinfo else tz.localize(check_in_time)
                lines.append(f"  - {name}: {local_time.strftime('%H:%M')} (muộn {late_minutes} phút)")
            lines.append("")

        if report.absent_users:
            lines.append("CHƯA CHECK-IN:")
            for name in report.absent_users:
                lines.append(f"  - {name}")
            lines.append("")

        if report.present_users:
            lines.append("ĐÃ CHECK-IN:")
            for name, check_in_time, is_late in report.present_users:
                local_time = check_in_time.astimezone(tz) if check_in_time.tzinfo else tz.localize(check_in_time)
                status = " (muộn)" if is_late else ""
                lines.append(f"  - {name}: {local_time.strftime('%H:%M')}{status}")

        return "\n".join(lines)

    @staticmethod
    def _get_monthly_points(year: int, month: int) -> List[MonthlyPointRow]:
        """
        Lấy điểm tháng của từng user, phân loại theo nguồn điểm.
        """
        with get_db_session() as session:
            active_users = (
                session.query(User)
                .filter(User.status == UserStatus.ACTIVE)
                .order_by(User.full_name)
                .all()
            )
            
            point_sums = (
                session.query(
                    PointLog.user_id,
                    PointLog.source_type,
                    func.sum(PointLog.points).label("points"),
                )
                .filter(
                    PointLog.month == month,
                    PointLog.year == year,
                )
                .group_by(PointLog.user_id, PointLog.source_type)
                .all()
            )
            
            # Map user_id -> source_type -> points
            point_map: Dict[int, Dict[str, int]] = {}
            for user_id, source_type, points in point_sums:
                if user_id not in point_map:
                    point_map[user_id] = {}
                point_map[user_id][source_type] = points or 0
            
            rows: List[MonthlyPointRow] = []
            for user in active_users:
                src_points = point_map.get(user.user_id, {})
                meeting = src_points.get("meeting", 0)
                evidence = src_points.get("evidence", 0)
                penalty = src_points.get("penalty", 0)
                absence = src_points.get("absence", 0)
                
                other = 0
                for src, val in src_points.items():
                    if src not in {"meeting", "evidence", "penalty", "absence"}:
                        other += val
                
                total = meeting + evidence + penalty + absence + other
                
                rows.append(
                    MonthlyPointRow(
                        user_id=user.user_id,
                        full_name=user.full_name,
                        total_points=total,
                        meeting_points=meeting,
                        evidence_points=evidence,
                        penalty_points=penalty,
                        absence_points=absence,
                        other_points=other,
                    )
                )
            
            # Sắp xếp theo tổng điểm giảm dần
            rows.sort(key=lambda r: r.total_points, reverse=True)
            return rows

    @staticmethod
    def generate_monthly_excel(year: int, month: int) -> io.BytesIO:
        """
        Generate monthly Excel report focusing on points (không tính đi muộn).
        Columns: User, Tổng điểm, Meeting, Evidence, Penalty, Absence, Khác.
        """
        point_rows = ExportService._get_monthly_points(year, month)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Điểm tháng"
        
        # Title
        ws.merge_cells("A1:H1")
        ws["A1"] = f"ĐIỂM THÁNG {month:02d}/{year}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        
        headers = [
            "STT",
            "Họ và tên",
            "Tổng điểm",
            "Meeting",
            "Evidence",
            "Penalty",
            "Absence",
            "Khác",
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = ExportService.HEADER_FILL
            cell.font = ExportService.HEADER_FONT
            cell.border = ExportService.THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        
        total_points = total_meeting = total_evidence = 0
        total_penalty = total_absence = total_other = 0
        
        for idx, row_data in enumerate(point_rows, 1):
            r = idx + 3
            ws.cell(row=r, column=1, value=idx).border = ExportService.THIN_BORDER
            ws.cell(row=r, column=2, value=row_data.full_name).border = ExportService.THIN_BORDER
            ws.cell(row=r, column=3, value=row_data.total_points).border = ExportService.THIN_BORDER
            ws.cell(row=r, column=4, value=row_data.meeting_points).border = ExportService.THIN_BORDER
            ws.cell(row=r, column=5, value=row_data.evidence_points).border = ExportService.THIN_BORDER
            ws.cell(row=r, column=6, value=row_data.penalty_points).border = ExportService.THIN_BORDER
            ws.cell(row=r, column=7, value=row_data.absence_points).border = ExportService.THIN_BORDER
            ws.cell(row=r, column=8, value=row_data.other_points).border = ExportService.THIN_BORDER
            
            total_points += row_data.total_points
            total_meeting += row_data.meeting_points
            total_evidence += row_data.evidence_points
            total_penalty += row_data.penalty_points
            total_absence += row_data.absence_points
            total_other += row_data.other_points
        
        total_row = len(point_rows) + 4
        ws.cell(row=total_row, column=1, value="Tổng").font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = ExportService.THIN_BORDER
        ws.cell(row=total_row, column=3, value=total_points).border = ExportService.THIN_BORDER
        ws.cell(row=total_row, column=4, value=total_meeting).border = ExportService.THIN_BORDER
        ws.cell(row=total_row, column=5, value=total_evidence).border = ExportService.THIN_BORDER
        ws.cell(row=total_row, column=6, value=total_penalty).border = ExportService.THIN_BORDER
        ws.cell(row=total_row, column=7, value=total_absence).border = ExportService.THIN_BORDER
        ws.cell(row=total_row, column=8, value=total_other).border = ExportService.THIN_BORDER
        
        # Widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 28
        for col_letter in ["C", "D", "E", "F", "G", "H"]:
            ws.column_dimensions[col_letter].width = 14
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_csv_report(year: int, month: int) -> str:
        """
        Generate monthly CSV report focusing on points.
        """
        point_rows = ExportService._get_monthly_points(year, month)
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow([f"Điểm tháng {month:02d}/{year}"])
        writer.writerow([])
        writer.writerow(["STT", "Họ và tên", "Tổng điểm", "Meeting", "Evidence", "Penalty", "Absence", "Khác"])
        
        total_points = total_meeting = total_evidence = 0
        total_penalty = total_absence = total_other = 0
        
        for idx, row_data in enumerate(point_rows, 1):
            writer.writerow([
                idx,
                row_data.full_name,
                row_data.total_points,
                row_data.meeting_points,
                row_data.evidence_points,
                row_data.penalty_points,
                row_data.absence_points,
                row_data.other_points,
            ])
            total_points += row_data.total_points
            total_meeting += row_data.meeting_points
            total_evidence += row_data.evidence_points
            total_penalty += row_data.penalty_points
            total_absence += row_data.absence_points
            total_other += row_data.other_points
        
        writer.writerow([])
        writer.writerow([
            "Tổng",
            "",
            total_points,
            total_meeting,
            total_evidence,
            total_penalty,
            total_absence,
            total_other,
        ])
        
        return output.getvalue()
